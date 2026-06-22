from urllib import request
from datetime import datetime, timedelta
import csv
import io
import openpyxl

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q, Sum, Count, F
from django.utils import timezone
from django.contrib import messages
import requests
from django.conf import settings
from django.core.paginator import Paginator
from django.http import HttpResponse

# ── IMPORTACIONES DE MODELOS ─────────────────────────────────────────────
from CuentasApp.models import Usuario, Secretaria, Rol, Estado
from MedicoApp.models import Medico, Disponibilidad
from PacienteApp.models import Paciente
from CitaApp.models import Cita 
from FacturacionApp.models import Pago
from InventarioApp.models import Producto

# ── IMPORTACIONES DE FORMULARIOS ─────────────────────────────────────────
from .forms import PQRSForm 
from CuentasApp.forms import (
    RegistroForm, 
    RegistroPacienteForm, 
    EditarPacienteForm,
    EditarPerfilPacienteForm
)

# =========================================================================
# --- VISTAS PÚBLICAS ---
# =========================================================================

def home(request):
    """Vista principal de la aplicación"""
    return render(request, 'Webapp/index.html')


def _enviar_correo_resend(asunto, cuerpo, destinatario, timeout=10, html=False):
    """
    Envía un correo a través de la API HTTP de Resend.

    Devuelve (ok: bool, error_msg: str|None).

    Parámetros:
      asunto, cuerpo, destinatario: lo evidente.
      timeout: segundos de espera para la llamada HTTP.
      html: si True, el cuerpo se manda como HTML; si False (default),
            como texto plano.

    Por qué no usamos send_mail() de Django con SMTP:
      Railway bloquea el tráfico SMTP saliente (puertos 25, 465, 587).
      Resend funciona sobre HTTPS (443), que Railway sí permite.

    LIMITACIÓN del plan gratuito de Resend sin dominio verificado:
      Solo se pueden enviar correos al email con el que se registró la
      cuenta de Resend. Si RESEND_TO_OVERRIDE está configurado en las
      variables de entorno, todos los correos se redirigen a ese email
      (útil en modo sandbox/pruebas). Cuando verifiques un dominio en
      Resend, elimina esa variable y los correos llegarán al destinatario
      real.
    """
    api_key = settings.RESEND_API_KEY
    if not api_key:
        return False, "RESEND_API_KEY no está configurada en variables de entorno."

    # En modo sandbox (sin dominio verificado en Resend), redirige todos
    # los correos al email autorizado para pruebas. Aplica para PQRS,
    # recordatorios, recuperación de contraseña — cualquier correo que
    # pase por esta función. Cuando tengas dominio verificado, elimina
    # la variable RESEND_TO_OVERRIDE de Railway y cada correo llegará
    # a su destinatario real.
    override = getattr(settings, 'RESEND_TO_OVERRIDE', None)
    if override:
        destinatario = override

    payload = {
        'from': settings.RESEND_FROM,
        'to': [destinatario],
        'subject': asunto,
    }
    if html:
        payload['html'] = cuerpo
    else:
        payload['text'] = cuerpo

    try:
        response = requests.post(
            'https://api.resend.com/emails',
            headers={
                'Authorization': f'Bearer {api_key}',
                'Content-Type': 'application/json',
            },
            json=payload,
            timeout=timeout,
        )
        if response.status_code in (200, 201):
            return True, None
        try:
            error_data = response.json()
            error_msg = error_data.get('message') or error_data.get('error') or response.text
        except Exception:
            error_msg = response.text
        return False, f"Resend respondió {response.status_code}: {error_msg}"
    except requests.exceptions.Timeout:
        return False, "La API de Resend tardó demasiado en responder."
    except requests.exceptions.RequestException as e:
        return False, f"Error de red al contactar Resend: {e}"


def contacto_pqrs(request):
    """Gestión de PQRS con envío automatizado de correos vía Resend."""
    if request.method == 'POST':
        form = PQRSForm(request.POST, user=request.user)
        if form.is_valid():
            if request.user.is_authenticated:
                nombre = f"{request.user.nombre} {request.user.apellidos}"
                email_usuario = request.user.correo
            else:
                nombre = form.cleaned_data.get('nombre')
                email_usuario = form.cleaned_data.get('email')

            tipo = form.cleaned_data['tipo']
            mensaje = form.cleaned_data['mensaje']

            asunto_clinica = f"NUEVA {tipo.upper()} - {nombre}"
            cuerpo_clinica = (
                f"Se ha recibido una solicitud:\n\n"
                f"Nombre: {nombre}\nCorreo: {email_usuario}\nTipo: {tipo}\n\n"
                f"Mensaje:\n{mensaje}"
            )

            asunto_usuario = f"Copia de su {tipo} - OdontoClinick"
            cuerpo_usuario = (
                f"Hola {nombre},\n\nHemos recibido tu {tipo.lower()} con éxito. "
                f"Pronto nos comunicaremos contigo.\n\nDetalles:\n\"{mensaje}\""
            )

            ok_clinica, err_clinica = _enviar_correo_resend(
                asunto_clinica, cuerpo_clinica, 'odontoclinick77@gmail.com'
            )

            ok_usuario = True
            err_usuario = None
            if email_usuario:
                ok_usuario, err_usuario = _enviar_correo_resend(
                    asunto_usuario, cuerpo_usuario, email_usuario
                )

            if ok_clinica and ok_usuario:
                messages.success(request, "✅ PQRS enviada con éxito. Revisa tu correo para ver la copia.")
                return redirect('home')
            elif ok_clinica and not ok_usuario:
                messages.warning(
                    request,
                    f"✅ Tu PQRS llegó a la clínica, pero no pudimos enviarte la copia ({err_usuario})."
                )
                return redirect('home')
            else:
                messages.error(request, f"❌ No se pudo enviar la PQRS a la clínica: {err_clinica}")
    else:
        form = PQRSForm(user=request.user)

    return render(request, 'Webapp/pqrs.html', {'form': form})


# =========================================================================
# --- VISTAS PRIVADAS (GESTIÓN Y DASHBOARDS) ---
# =========================================================================

@login_required
def panel_secretaria(request):
    """Dashboard para el rol de Secretaria con Sala de Espera"""
    if request.user.id_rol.nombre_rol not in ['Secretaria', 'Administrador']:
        return redirect('home')
        
    total_pacientes = Usuario.objects.filter(id_rol__nombre_rol='Paciente').count()
    hoy = timezone.now().date()
    
    citas_hoy_count = Cita.objects.filter(
        fecha_hora__date=hoy
    ).exclude(
        id_estado_cita__nombre_estado__icontains='Cancelada'
    ).count()

    pacientes_espera = Cita.objects.filter(
        id_estado_cita__nombre_estado__icontains='En Espera',
        fecha_hora__date=hoy
    ).select_related(
        'id_paciente__id_usuario', 
        'id_doctor__id_usuario',
        'id_estado_cita'
    ).order_by('hora_llegada')
    
    contexto = {
        'total_pacientes': total_pacientes,
        'citas_hoy_count': citas_hoy_count,
        'nombre_usuario': request.user.nombre,
        'pacientes_espera': pacientes_espera,
    }
    
    return render(request, 'Webapp/panel_secretaria.html', contexto)


@login_required
def registro_integral_paciente(request):
    """Registro transaccional unificado de Usuario y Paciente"""
    if request.user.id_rol.nombre_rol not in ['Secretaria', 'Administrador']:
        return redirect('home')
    if request.method == 'POST':
        form_user = RegistroForm(request.POST)
        form_paciente = RegistroPacienteForm(request.POST)
        if form_user.is_valid() and form_paciente.is_valid():
            try:
                with transaction.atomic():
                    nuevo_usuario = form_user.save(commit=False)
                    nuevo_usuario.id_rol = Rol.objects.get(nombre_rol='Paciente')
                    nuevo_usuario.id_estado = Estado.objects.filter(nombre_estado__icontains='Activo').first() or Estado.objects.get(id_estado=1)
                    nuevo_usuario.save() 
                    
                    paciente_instancia, _ = Paciente.objects.get_or_create(id_usuario=nuevo_usuario)
                    form_p_final = RegistroPacienteForm(request.POST, instance=paciente_instancia)
                    if form_p_final.is_valid():
                        form_p_final.save()
                    
                    messages.success(request, f"Paciente {nuevo_usuario.nombre} registrado correctamente.")
                    return redirect('lista_pacientes')
            except Exception as e:
                messages.error(request, f"Error: {e}")
    else:
        form_user = RegistroForm()
        form_paciente = RegistroPacienteForm()
    return render(request, 'Webapp/registrar_paciente.html', {'form_user': form_user, 'form_paciente': form_paciente})


@login_required
def lista_pacientes(request):
    """Listado paginado de pacientes con filtros de búsqueda"""
    if request.user.id_rol.nombre_rol not in ['Secretaria', 'Administrador']:
        return redirect('home')
    
    query = request.GET.get('q')
    pacientes_list = Usuario.objects.filter(id_rol__nombre_rol='Paciente').order_by('nombre')
    
    if query:
        pacientes_list = pacientes_list.filter(
            Q(nombre_usuario__icontains=query) | Q(nombre__icontains=query) | Q(apellidos__icontains=query)
        )
    
    paginator = Paginator(pacientes_list, 10)
    page_number = request.GET.get('page')
    pacientes = paginator.get_page(page_number)
    
    return render(request, 'Webapp/lista_pacientes.html', {'pacientes': pacientes, 'query': query})


@login_required
def editar_paciente(request, id_usuario):
    """Edición mediante QuerySet.update() para prevenir colisiones en los métodos de guardado"""
    if request.user.id_rol.nombre_rol not in ['Secretaria', 'Administrador']:
        return redirect('home')

    usuario_instancia = get_object_or_404(Usuario, id_usuario=id_usuario)
    paciente_instancia = get_object_or_404(Paciente, id_usuario=usuario_instancia)
    
    if request.method == 'POST':
        form_user = EditarPacienteForm(request.POST, instance=usuario_instancia)
        form_clinico = RegistroPacienteForm(request.POST, instance=paciente_instancia)
        
        if 'id_estado' in form_user.fields: form_user.fields['id_estado'].required = False
        if 'id_rol' in form_user.fields: form_user.fields['id_rol'].required = False

        if form_user.is_valid() and form_clinico.is_valid():
            try:
                with transaction.atomic():
                    datos_usuario = form_user.cleaned_data
                    
                    Usuario.objects.filter(id_usuario=id_usuario).update(
                        nombre=datos_usuario.get('nombre', usuario_instancia.nombre),
                        apellidos=datos_usuario.get('apellidos', usuario_instancia.apellidos),
                        nombre_usuario=datos_usuario.get('nombre_usuario', usuario_instancia.nombre_usuario),
                        correo=datos_usuario.get('correo', usuario_instancia.correo),
                        telefono=datos_usuario.get('telefono', usuario_instancia.telefono),
                    )
                    form_clinico.save()
                    
                messages.success(request, f"¡Paciente {usuario_instancia.nombre} actualizado!")
                return redirect('lista_pacientes')
            except Exception as e:
                messages.error(request, f"Error técnico: {e}")
        else:
            messages.error(request, "Error en los datos. Revisa el formulario.")
    else:
        form_user = EditarPacienteForm(instance=usuario_instancia)
        form_clinico = RegistroPacienteForm(instance=paciente_instancia)
    
    return render(request, 'Webapp/editar_paciente.html', {
        'form_user': form_user, 'form_clinico': form_clinico, 'paciente': usuario_instancia
    })


@login_required
def carga_masiva_pacientes(request):
    """Carga masiva por filas usando transacciones atómicas individuales"""
    if request.user.id_rol.nombre_rol not in ['Secretaria', 'Administrador']:
        return redirect('home')

    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        archivo = request.FILES['archivo_excel']
        
        try:
            wb = openpyxl.load_workbook(archivo)
            hoja = wb.active
            creados = 0
            errores = 0

            rol_paciente = Rol.objects.get(nombre_rol='Paciente')
            estado_activo = Estado.objects.filter(nombre_estado__icontains='Acti').first()

            for fila in hoja.iter_rows(min_row=2, values_only=True):
                user_val = str(fila[0]) if fila[0] else None
                if not user_val: 
                    continue 

                try:
                    with transaction.atomic():
                        nuevo_u, created = Usuario.objects.update_or_create(
                            nombre_usuario=user_val,
                            defaults={
                                'nombre': fila[2],
                                'apellidos': fila[3],
                                'correo': fila[4],
                                'telefono': fila[5],
                                'id_rol': rol_paciente,
                                'id_estado': estado_activo,
                            }
                        )
                        
                        if created:
                            nuevo_u.set_password(str(fila[1]))
                            nuevo_u.save()

                        Paciente.objects.update_or_create(
                            id_usuario=nuevo_u,
                            defaults={
                                'fecha_nacimiento': fila[6],
                                'direccion': fila[7],
                                'eps': fila[8],
                                'rh': fila[9],
                                'alergias': fila[10],
                                'enfermedades_preexistentes': fila[11],
                                'contacto_emergencia_nombre': fila[12],
                                'contacto_emergencia_telefono': fila[13]
                            }
                        )
                        creados += 1
                        
                except Exception as e:
                    print(f"Error procesando fila {user_val}: {e}")
                    errores += 1

            messages.success(request, f"Proceso finalizado. Procesados con éxito: {creados}. Errores: {errores}")
            return redirect('lista_pacientes')
        except Exception as e:
            messages.error(request, f"Error crítico al leer el archivo: {e}")

    return render(request, 'Webapp/carga_masiva.html')


@login_required
def detalle_paciente(request, id_usuario):
    """Vista detallada del historial clínico resumido de un paciente"""
    if request.user.id_rol.nombre_rol not in ['Secretaria', 'Administrador']:
        return redirect('home')
    usuario = get_object_or_404(Usuario, id_usuario=id_usuario)
    paciente_clinico = get_object_or_404(Paciente, id_usuario=usuario)
    citas_recientes = Cita.objects.filter(id_paciente=paciente_clinico).order_by('-fecha_hora')[:5]
    return render(request, 'Webapp/detalle_paciente.html', {'u': usuario, 'p': paciente_clinico, 'citas': citas_recientes})


@login_required
def descargar_plantilla_pacientes(request):
    """Generación dinámica del Excel guía para cargas masivas"""
    wb = openpyxl.Workbook()
    hoja = wb.active
    hoja.title = "Plantilla_Pacientes"

    encabezados = [
        'nombre_usuario', 'password', 'nombre', 'apellidos', 
        'correo', 'telefono', 'fecha_nacimiento', 'direccion', 
        'eps', 'rh', 'alergias', 'enfermedades_preexistentes', 
        'contacto_emergencia_nombre', 'contacto_emergencia_telefono'
    ]
    hoja.append(encabezados)

    ejemplo = [
        '10102020', 'Pass123*', 'Juan', 'Perez', 
        'juan.perez@email.com', '3001234567', '1995-10-25', 'Calle 123', 
        'Sura', 'O+', 'Ninguna', 'Ninguna', 
        'Maria Perez', '3109876543'
    ]
    hoja.append(ejemplo)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=plantilla_pacientes_odontoclinick.xlsx'
    wb.save(response)
    return response